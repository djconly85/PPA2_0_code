# --------------------------------
# Name: utils.py
# Purpose: Provides general PPA functions that are used throughout various PPA scripts and are not specific to any one PPA script
#
#
# Author: Darren Conly
# Last Updated: <date>
# Updated by: <name>
# Copyright:   (c) SACOG
# Python Version: 3.x
# --------------------------------
# import win32com.client
# import gc # python garbage collector
import os
import datetime as dt


import openpyxl
from openpyxl.drawing.image import Image
import xlwings as xw
import pandas as pd
import arcpy

import ppa_input_params as p


def make_fl_conditional(fc, fl):
    '''check if a feature layer name exists; if it does, delete feature layer and remake it.
    purpose is to ensure the feature layer name corresponds to the correct feature class.'''
    if arcpy.Exists(fl):
        arcpy.Delete_management(fl)
    arcpy.MakeFeatureLayer_management(fc, fl)
    
    
def esri_field_exists(in_tbl, field_name):
    fields = [f.name for f in arcpy.ListFields(in_tbl)]
    if field_name in fields:
        return True
    else:
        return False


def esri_object_to_df(in_esri_obj, esri_obj_fields, index_field=None):
    '''converts esri gdb table, feature class, feature layer, or SHP to pandas dataframe'''
    data_rows = []
    with arcpy.da.SearchCursor(in_esri_obj, esri_obj_fields) as cur:
        for row in cur:
            out_row = list(row)
            data_rows.append(out_row)

    out_df = pd.DataFrame(data_rows, index=index_field, columns=esri_obj_fields)
    return out_df


def rename_dict_keys(dict_in, new_key_dict):
    '''if dict in = {0:1} and dict out supposed to be {'zero':1}, this function renames the key accordingly per
    the new_key_dict (which for this example would be {0:'zero'}'''
    dict_out = {}
    for k, v in new_key_dict.items():
        if k in list(dict_in.keys()):
            dict_out[v] = dict_in[k]
        else:
            dict_out[v] = 0
    return dict_out



def join_xl_import_template(template_xlsx, template_sheet, in_df):
    '''takes in import tab of destination Excel sheet, then left joins to desired output dataframe to ensure that
    output CSV has same rows every time, even if data frame that you're joining doesn't
    have all records'''
    df_template = pd.read_excel(template_xlsx, template_sheet)
    df_template = pd.DataFrame(df_template[df_template.columns[0]]) # get rid of all columns except for data items column
    df_template = df_template.set_index(df_template.columns[0]) # set data items column to be the index
    
    df_out = df_template.join(in_df) # left join the df from import sheet template to the df with data based on index values
    
    return df_out


class Publish(object):
    def __init__(self, in_df, xl_template, import_tab, xl_out, xlsheets_to_pdf=None, map_img_dict=None, 
                 proj_name='UnnamedProject'):
        self.in_df = in_df
        self.import_tab = import_tab
        self.xl_out = xl_out
        self.map_img_dict = map_img_dict  # {<image file name>: [<sheet to put image on>, <row>, <col>]}
        self.xl_template = xl_template
        self.xl_workbook = openpyxl.load_workbook(self.xl_template)
        self.time_sufx = str(dt.datetime.now().strftime('%m%d%Y_%H%M'))
        
        self.xlsheets_to_pdf = xlsheets_to_pdf
        self.pdf_out = os.path.join(p.dir_pdf_output, '{}_{}.pdf'.format(proj_name, self.time_sufx ))

    def overwrite_df_to_xlsx(self, unused=0, start_row=0, start_col=0):  # why does there need to be an argument?
        '''Writes pandas dataframe <in_df_ to <tab_name> sheet of <xlsx_template> excel workbook.'''
        in_df = self.in_df.reset_index()
        df_records = in_df.to_records(index=False)
        
        # get header row for output
        out_header_list = [list(in_df.columns)]  # get header row for output
        
        out_data_list = [list(i) for i in df_records]  # get output data rows
    
        comb_out_list = out_header_list + out_data_list
    
        ws = self.xl_workbook[self.import_tab]
        for i, row in enumerate(comb_out_list):
            for j, val in enumerate(row):
                cell = ws.cell(row=(start_row + (i + 1)), column=(start_col + (j + 1)))
                if (cell):
                    cell.value = val
    
    def insert_image_xlsx(self, sheet_name, rownum, col_letter, img_file):
        '''inserts image into specified sheet and cell within Excel workbook'''
        ws = self.xl_workbook[sheet_name]
        cell = '{}{}'.format(col_letter, rownum) # will be where upper left corner of image placed
        img_obj = Image(img_file)
        ws.add_image(img_obj, cell)
        
    def make_new_excel(self):
        '''takes excel template > writes new values to import/updates charts, then inserts indicated images at specified locations'''
        self.overwrite_df_to_xlsx(self) # write data to import tab
        
        # insert map images at appropriate locations
        if self.map_img_dict:
            for imgfile, location_info in self.map_img_dict.items():
                sheetname = location_info[0]
                imgrow = location_info[1]
                imgcol = location_info[2]
                self.insert_image_xlsx(sheetname, imgrow, imgcol, imgfile)
            
        self.xl_workbook.save(self.xl_out)
        
    def make_pdf(self):
        arcpy.AddMessage("Publishing to PDF...")
        # self.make_new_excel(self)
        
        # 1/16/2020 - waiting for Z to make stable, online-friendly version of this function
        xw.App.visible = False
        
        if not os.path.exists(self.xl_out):
            self.make_new_excel(self)
        
        wb = xw.Book(self.xl_out)
            
        for s in self.xlsheets_to_pdf:
            out_sheet = wb.sheets[s]
            pdf_out = os.path.join(p.dir_pdf_output, '{}{}_{}.pdf'.format(proj_name, s, self.time_sufx))
            out_sheet.api.ExportAsFixedFormat(0, pdf_out)



"""
def insert_image_xlsx(xl_workbook, sheet_name, rownum, col_letter, img_file):
    ws = xl_workbook[sheet_name]
    cell = '{}{}'.format(col_letter, rownum) # will be where upper left corner of image placed
    img_obj = Image(img_file)
    ws.add_image(img_obj, cell)

def overwrite_df_to_xlsx(in_df, xl_workbook, tab_name, start_row=0, start_col=0):
    '''Writes pandas dataframe <in_df_ to <tab_name> sheet of <xlsx_template> excel workbook.'''

    in_df = in_df.reset_index()
    df_records = in_df.to_records(index=False)
    
    # get header row for output
    out_header_list = [list(in_df.columns)]  # get header row for output
    
    out_data_list = [list(i) for i in df_records]  # get output data rows

    comb_out_list = out_header_list + out_data_list

    ws = xl_workbook[tab_name]
    for i, row in enumerate(comb_out_list):
        for j, val in enumerate(row):
            cell = ws.cell(row=(start_row + (i + 1)), column=(start_col + (j + 1)))
            if (cell):
                cell.value = val
"""



    
"""  
def excel2pdf(in_xlsx, out_pdf, sheets_to_pdf_list):
    '''converts excel to PDF file.
    sheets_to_pdf_list = list of the sheets in the Excel workbook you want to include in output PDF
    WARNING - THIS FUNCTION WILL NOT WORK WHEN PUBLISHED AS AN ESRI ONLINE TOOL. ONLY WORKS ON DESKTOP.
    '''
    excel_app_obj = win32com.client.Dispatch("Excel.Application")  # creates object for interacting with Excel files
    
    # Settings so that the targeted Excel template workbook does not visibly open when you activate it.
    excel_app_obj.Visible = False 
    # excel_app_obj.ScreenUpdating = False
    # excel_app_obj.DisplayAlerts = False
    # excel_app_obj.EnableEvents = False

    wb = excel_app_obj.Workbooks.Open(in_xlsx)  # open a specific Excel workbook (XLSX file)
    
    # indicate which sheets you want to open
    ws_index_list = sheets_to_pdf_list  # ws_index_list = [2, 3]  # say you want to print these sheets

    wb.Worksheets(ws_index_list).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, out_pdf)
    
    # close/clean up so that, if you loop through projects, you don't get any errors as a result of unfinished processes.
    excel_app_obj.Quit()
    gc.collect()
"""
    
    
    
    
    
    
"""
def join_csv_template(template_csv, in_df):
    '''takes in template CSV, left joins to desired output dataframe, ensure that
    output CSV has same rows every time, even if data frame that you're joining doesn't
    have all records'''
    df_template = pd.read_csv(template_csv)
    df_template = df_template.set_index(df_template.columns[0])
    
    df_out = df_template.join(in_df)
    
    return df_out
"""

