# -*- coding: utf-8 -*-
"""
Created on Tue Mar 31 15:51:58 2020

@author: dconly
"""

import os

from openpyxl import load_workbook

def move_sheets(wb, sheets_to_move):
    
    sheet_objs = wb._sheets
    # title_posn = wb.sheetnames.index('0ATitlePg')
    disclaimer_posn = wb.sheetnames.index('0BUsingThisReport')
    
    for i, sheet in enumerate(sheets_to_move):
        start_pos = wb.sheetnames.index(sheet)
        posns_to_move = i + 1
        destination = disclaimer_posn + posns_to_move
        
        sheet_obj2move = sheet_objs.pop(start_pos) # cut sheet out of original position
        sheet_objs.insert(destination, sheet_obj2move) # paste into desired position

def color_sheets(wb, sheet_names):
    for sheet_name in sheet_names:
        sheet_posn = wb.sheetnames.index(sheet_name)
        sheet_obj = wb._sheets[sheet_posn]
        sheet_obj.sheet_properties.tabColor = "45b045" # RGB color code
    



if __name__ == '__main__':
    wkng_dir = r'C:\Users\dconly\Desktop\TempLocalPPA\Test outputs'
    in_xlsx = os.path.join(wkng_dir, "JH_PPA_test3031920201457.xlsx")
    out_xlsx = os.path.join(wkng_dir, "output_test.xlsx")
    
    
    wb_in = load_workbook(in_xlsx)
    perf_outcomes = ['2ReduceCongestion','4EconProsperity','5Freight']
    
    

    move_sheets(wb_in, perf_outcomes)
    
    sheets_to_color = ['0ATitlePg', '0BUsingThisReport'] + perf_outcomes
    
    color_sheets(wb_in, sheets_to_color)
    
    wb_in.save(out_xlsx)
    wb_in.close()
    
    
'''
Desired process:
    1 - have list of user-selected outcomes
    2 - for each user-selected outcome, pop it from its original position
    '''