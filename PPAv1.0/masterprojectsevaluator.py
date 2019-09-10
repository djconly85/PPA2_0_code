'''
Purpose - create tabular data table with PPA projects as rows and metrics as columns.
    Metrics are in excel sheet Q:\ProjectLevelPerformanceAssessment\DataLayers_Proof_of_Concept\Spreadsheet\RawDataTableLayout.xlsx
    
'''
#..changes the order of import so that more generic ones goes on the top.
import sys
import re
import os
import datetime
import time
import csv
import shutil
import openpyxl
import arcpy
import ppautils 

#..Note, all these are hardcoded layers, which should be replaced by a Dataset + filter.
#..i.e., the dataset is a constant pointing to a featureclass, and the 

LN_centerLine = "Region_Centerline2017fwytag2"
LN_AllParcel = "parcel_all"
LN_JobCenter = "JobCenters"
LN_BikewaysCurrent = "AllBikeways2015"
LN_BikewaysFuture = "ProposedBikeways2015"
LN_TransitData = "TrnStopLocnWTrpCnt11172017"
LN_TIMS = "Collisions2011to2016fwytag"
LN_TAZ = "TAZ07_01"
LN_Community_Type = "community_type"
FN_F2016_Proje = "F2016_Proje" 

FNH_CommunityType = "Proj_Community_Type"    #HEADER for line overlay community type, value comes projectpolyline overlaying with LN_Community_Type's type field (in the overlaying XLine, named type_1, if polyline also has a field named TYPE

C_XLSX_Template = "Digest_Summary_Template.xlsx"
C_MissingValue = -999

C_OK = 'OK'

def trace():
    import traceback, inspect
    tb = sys.exc_info()[2]
    tbinfo = traceback.format_tb(tb)[0]
    # script name + line number
    line = tbinfo.split(", ")[1]
    filename = inspect.getfile(inspect.currentframe())
    # Get Python syntax error
    synerror = traceback.format_exc().splitlines()[-1]
    return line, filename, synerror #return line number, name of file with error line, and type of error

def GetThisFileName():
    import inspect
    return inspect.getfile(inspect.currentframe())

def GetThisDir():
    import inspect 
    return os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))            

def writecsv2xlsx(txtFile, xlsxFile, outxlsxFile, start_row=4, start_col=0):
	try:
		is_number = re.compile(r'\d+(?:,\d*)?')
		wb = openpyxl.load_workbook(xlsxFile)
		ws = wb['1-import']
		with open(txtFile, 'r') as infile:
			for i, line in enumerate(infile):
				#..Skip first line:
				if(i>0): 
					lValues = line.split(",") 
					for j, val in enumerate(lValues):
						if is_number.match(val):
							val = float(val)
							cell = ws.cell(row = (start_row + i), column = (start_col + (j+1)))
							if(cell):
								cell.value = val
						else:
							cell = ws.cell(row = (start_row + i), column = (start_col + (j+1)))
							if(cell):
								cell.value = val 
        #ws.write_string(start_row + i, start_col + (j+1), val)
		wb.save(outxlsxFile)     
	except:
		sMsg = "{}".format(trace()) 
		print(sMsg) 


#..Note, if you'd like to showTime the processing time of dowork function, uncomment the following @ppautils.showTime line.
#@ppautils.showTime
def dowork(input_project_lines, buffer_dist_1, buffer_dist_2, buffer_dist_jc, workspace, project_type, dYearDatasets, in_year, dProperties, pScratchGDB, debugLevel=0):

    pyFile = GetThisFileName()
    pyFilePath = os.path.dirname(pyFile) 
    arcpy.AddMessage("pyFile={}\npyFilePath={}".format(pyFile, pyFilePath))
     
    arcpy.env.overwriteOutput = True
    tReturn = ()                        #.. Holds the returning values, (sOK, ...) 
    pScratchworkspace = "in_memory"     # arcpy.env.scratchWorkspace
    dt = time.clock()                   #..used to gage the processing time.
    try:
        project_type_orig = ""
        with arcpy.da.SearchCursor(input_project_lines, [FN_F2016_Proje]) as rows:
            for row in rows:
                project_type_orig = row[0]
       
        if(project_type!=''):
            try:
                arcpy.CalculateField_management(input_project_lines, FN_F2016_Proje, "\'{}\'".format(project_type), "PYTHON_9.3")
            except:
                sMsg = "{}, {}".format(arcpy.GetMessages(2), trace())
                arcpy.AddMessage(sMsg) 

        if(debugLevel & 2)==2: arcpy.AddMessage("F2016_Proje ORG={} New={}".format(project_type_orig, project_type))  

        dateSuffix = str(datetime.date.today().strftime('%m%d%Y'))
        #===============USER-DEFINED PARAMETERS=========================
        # input choices
        #Year = 2036        #year to use for ILUT and model network data
        AllItems = True    #False   #True #True/False. Whether you want to summarize all projects or just select projects

        # Specify directories
        #workSpace = r"D:\10Data\SACOG_PPA\PPA Files4ESRI_01312018\PPA_layers_4ESRI.gdb"      #'Q:\ProjectLevelPerformanceAssessment\DataLayers_Proof_of_Concept\PPA_layers.gdb'
        arcpy.env.workspace = workspace
		
		#used for debugging, check to see if all feature datasets and feature classes exist in workspace
        #arcpy.AddMessage("Layers: {}".format(arcpy.ListFeatureClasses()))
        #arcpy.AddMessage("Datasets: {}".format(arcpy.ListDatasets()))

		
        accessibilityTxtDir = os.path.dirname(workspace)           # os.path.dirname(workspace)     #  r"D:\10Data\SACOG_PPA\PPA Files4ESRI_01312018"                 # r'Q:\ProjectLevelPerformanceAssessment\DataLayers_Proof_of_Concept\Accessbility Metric'
        outTextDir = os.path.join(accessibilityTxtDir, "output")   #  r"D:\10Data\SACOG_PPA\PPA Files4ESRI_01312018\output"                   #    r'Q:\ProjectLevelPerformanceAssessment\DataLayers_Proof_of_Concept\Spreadsheet\masterTableOutputs'

        if(os.path.exists(outTextDir)==False): os.makedirs(outTextDir) 
            
        projectIdCol = "F2016_Proje"   #column with the unique project ids
        #data layer inputs

        FN_TYPE = "type" 
        fcCommunityType = os.path.join(workspace, LN_Community_Type) 
        #..if input project line contains "TYPE" field, then the TYPE field in the Community_type layer would be renamed in the intersect line (xline) to TYPE_1
        if(len(arcpy.ListFields(fcCommunityType, FN_TYPE))>0): FN_TYPE = "TYPE_1" 

        xLineName = "XLine{}".format(int(time.clock())) 
        fcXLine = os.path.join(pScratchGDB, xLineName) 
        inputLineMsg = "Layer: '{}' exists = {}".format(input_project_lines, arcpy.Exists(input_project_lines))
        arcpy.AddMessage(inputLineMsg)
        comTypesMsg = "Layer: '{}' exists = {}".format(fcCommunityType, arcpy.Exists(fcCommunityType))
        arcpy.AddMessage(comTypesMsg)
		
        arcpy.Intersect_analysis([input_project_lines, fcCommunityType], fcXLine, output_type = "LINE") #"{};{}".format(input_project_lines, fcCommunityType)
        dCommunityTypes = dict() 
        lengMax = 0 
        sTypeMax = ""
        sType = ""
        with arcpy.da.SearchCursor(fcXLine, ["SHAPE@", FN_TYPE]) as rows: #cursor for the intersected lines
            for row in rows:
                try:
                    pLine = row[0]
                    sType = row[1]
                    lengthis = pLine.length
                    if(sType in dCommunityTypes):
                        leng = dCommunityTypes[sType]
                        leng = leng + lengthis
                        dCommunityTypes[sType] = leng
                    else:
                        dCommunityTypes.setdefault(sType, lengthis)
                        leng = lengthis
                     
                    if(sTypeMax=="") or (leng>lengMax):
                        sTypeMax = sType 
                        lengMax = leng
                    
                except:
                    pass  
        dProperties.setdefault(FNH_CommunityType, sTypeMax)     
        if(debugLevel & 2)==2: arcpy.AddMessage("dCommunityTypes={}\nCommType={}".format(dCommunityTypes, sTypeMax))

        # specify list of projects if you are testing and don't want to summarize all projects.
        specProjList = ['PLA25548',
                        'CAL20432',
                        'SAC24704'] 


        start_time = time.time()
        #create output file
        #myfile = open(outputFile, 'w')
        #..dynamically constructed fc names. --- year can take in only the values of 2012, 2036 at this point.
        year = 0 
        try:
            year = int(in_year)
        except:
            year = -1

        if (year > 0):            
            lYears = [year]
        else:
            lYears = list(dYearDatasets.keys()) #dYearDatasets.keys()

        if(year>0):
            outputFileName = "output_projects{}_{}.csv".format(year, dateSuffix)
            outputFile = os.path.join(arcpy.env.scratchFolder,  outputFileName)    # os.path.join(outTextDir, outputFileName) 
        else:
            outputFileName = "output_projects_allyears_{}.csv".format(dateSuffix)
            outputFile = os.path.join(arcpy.env.scratchFolder,  outputFileName)    # os.path.join(outTextDir, outputFileName) 
             
        #outputFileName = "output_projects{}_{}.csv".format(year, dateSuffix)
        #outputFile = os.path.join(arcpy.env.scratchFolder,  outputFileName)    # os.path.join(outTextDir, outputFileName) 
        if(debugLevel & 1)==1:  arcpy.AddMessage("  lYears={}".format(lYears))
        #write header row to output file
        #==================DATA IMPORT AND PREPARATION========================
        #all-parcel file project stats
        acres_AG = "acres_AG"
        netACdu = "netACdu"
        netACemp = "netACemp"
        if(len(lYears)==1):
            sYear = str(lYears[0])
            yr2char = sYear[-2:]
            acres_AG = "{}_{}".format(acres_AG, yr2char) 
            netACdu = "{}_{}".format(netACdu, yr2char) 
            netACemp = "{}_{}".format(netACemp, yr2char)

        varListAllParcelHeader = ["acres", acres_AG,"du_BO","emp_BO",
                            netACdu, netACemp]
                    
        #ILUT
        varListILUT = ["DU_TOT","POP_TOT","POP_EJ","HH_TOT_P","EMPEDU","EMPFOOD","EMPGOV","EMPOFC",
                        "EMPOTH","EMPRET","EMPSVC","EMPMED","EMPIND","EMPTOT_S","STUD_K12","STUD_UNI",
                        "ENR_K12","ENR_UNI",
                        "VT_TOT","VMT_TOT","CVMT_TOT","PT_TOT_RES","SOV_TOT_RES","HOV_TOT_RES",
                        "TRN_TOT_RES","BK_TOT_RES","WK_TOT_RES","VT_TOT_cv","CV3_VT","VMT_TOT_cv",
                        "CV3_VMT","CVMT_TOT_cv","CV3_CVMT"]


        #intersections
        varListIntersecn = ["INTXN_1WAY","INTXN_3WAY","INTXN_4WAY"]
        # Job Center polygons

        varListjobCenter = ['JOB_CTR']

        #street centerline length
        varListCLineDist = ['BUF_CTRLINE_MI']
        # Bike Facilities
        varListBikeNtwk = ['BIKC1_MIBUFF','BIKC2_MIBUFF','BIKC3_MIBUFF','BIKC4_MIBUFF']

        # Google Transit
        varListTransit = ['TRAN_LOCNS','TRAN_EVENTS']
            #service density/vehicle stops (use pointSum function)
            #unique stop locations (look to UOP NPMRDS script for inspiration)

        # collisons
        varListCollisionSums = ['PEDKILL','PEDINJ','BICKILL','BICINJ'] 
        varListCollisionsCounts = ['FWY_COLLN',
                                    'NON_FWY_COLLN', 
                                    'FWY_FATAL_COLLN',
                                    'NONFWY_FATAL_COLLN'] #count of all collisions


        # Model network (buffer)
        varListModelBuffer = ['FwyVMT_buff','FwyCVMT_buff','FwyLnMi_buff','FwyRteMi_buff',
                                'StVMT_buff','StCVMT_buff','StLnMi_buff','StRteMi_buff']

        #TAZ-based accessibilityTxtDir
        varListAccessibility = ['JOBS30D_PAVG','JOBS45T_PAVG','WKR30D_JBAVG','WKR45T_JBAVG']
        lproperty_keys = list(dProperties.keys()) #dProperties.keys()
        lproperty_keys = sorted(lproperty_keys)  
        propertyKeys = ",".join("{}".format(key) for key in lproperty_keys) 
        propertyValues = ",".join("{}".format(str(dProperties[key])) for key in lproperty_keys)
        if(debugLevel & 1)==1: arcpy.AddMessage("{}\n{}".format(propertyKeys, propertyValues)) 
        if(os.path.exists(outputFile)):
            try:
                arcpy.Delete_management(outputFile) 
            except:
                sMsg = "  {}, {}".format(arcpy.GetMessages(2), trace())
                arcpy.AddWarning(sMsg) 

        with open(outputFile, 'w') as myfile:
            #=================PREPARE OUTPUT TABLE================================
            #header row elements are concatenated lists converted to strings with comma-separated words
            Header_Year = 'Year'
            Header_ProjInfo = 'PROJ_LEN_MI' #'projectID,PROJ_TYPE,commun_type,PROJ_LEN_MI'
            Header_buffArea= 'buffarea_qm,buffarea_hm'
            Header_allParcel = ','.join(varListAllParcelHeader)   #   (varListAllParcel) #makes list into comma-sep'd string
            Header_ILUT = ','.join(varListILUT)
            Header_Intersecn = ','.join(varListIntersecn)
            Header_jobCenter = ','.join(varListjobCenter)
            Header_CLineDist = ','.join(varListCLineDist)
            Header_bikeCLDist = ','.join(varListBikeNtwk)
            Header_Transit = ','.join(varListTransit)
            Header_TIMS = ','.join(varListCollisionsCounts) + ',' + ','.join(varListCollisionSums)
            Header_modelBuff = ','.join(varListModelBuffer)
            Header_access = ','.join(varListAccessibility)
            myfile.write(Header_Year + ','
                        + propertyKeys + ','
                        + Header_ProjInfo + ','
                        + Header_buffArea + ','
                        + Header_allParcel + ','
                        + Header_ILUT + ','
                        + Header_Intersecn + ','
                        + Header_jobCenter + ','
                        + Header_CLineDist + ','
                        + Header_bikeCLDist + ','
                        + Header_Transit + ','
                        + Header_TIMS + ','
                        + Header_modelBuff + ','
                        + Header_access + ','
                        + '\n') #write out headers

            #=======================make feature layers=========================
            #.. Replace all the "layernames" into fl_layername variables.  zye
            fl_inprojectslist = "inprojectsList_lyr"
            fl_inAllParcel = "inAllParcel_lyr"
            fl_jobCenter = "inJobCenter_lyr"
            fl_gtfs = "inGTFS_lyr"
            fl_tims = "inTIMS_lyr"
            fl_taz = "inTAZ_lyr"
            fl_centerline = "inCenterline_lyr"
            fl_inilut = "inILUT_lyr"
            fl_inModelNetwork = "inModelNetwork_lyr"
            fl_inInterSect = "inInterSect_lyr" 
            fl_bikeways = "bikeways_lyr"

            if(arcpy.Exists(fl_inprojectslist)): arcpy.Delete_management(fl_inprojectslist) 
            arcpy.MakeFeatureLayer_management(input_project_lines, fl_inprojectslist)   # "inprojectsList_lyr")
            arcpy.MakeFeatureLayer_management(LN_AllParcel, fl_inAllParcel)
            arcpy.MakeFeatureLayer_management(LN_JobCenter, fl_jobCenter)
            #arcpy.MakeFeatureLayer_management(LN_BikewaysCurrent,"currBikeways_lyr") #shouldn't be needed as of 3/2/2018. See code in "__main__" part
            #arcpy.MakeFeatureLayer_management(LN_BikewaysFuture,"futBikeways_lyr")
            arcpy.MakeFeatureLayer_management(LN_TransitData, fl_gtfs)
            arcpy.MakeFeatureLayer_management(LN_TIMS, fl_tims)
            arcpy.MakeFeatureLayer_management(LN_TAZ, fl_taz)
            arcpy.MakeFeatureLayer_management(LN_centerLine, fl_centerline)
            for yr in lYears:
                lDatasets = dYearDatasets[yr]
                parcel = lDatasets[0]       #ILUT
                intersect = lDatasets[1]    #interSect
                modelnet = lDatasets[2]     #modelNet
                accessibility_tbl = lDatasets[3] #TAZ accessbility table
                bikeways = lDatasets[4]     #bikeway layer

                arcpy.MakeFeatureLayer_management(parcel,fl_inilut)
                arcpy.MakeFeatureLayer_management(modelnet, fl_inModelNetwork)
                arcpy.MakeFeatureLayer_management(intersect,fl_inInterSect)
                arcpy.MakeFeatureLayer_management(bikeways,fl_bikeways)
                
                #accessTxtName = "access_{}.csv".format(yr)
                #accessTxt =  os.path.join(accessibilityTxtDir, accessTxtName)     

                #all-parcel file project stats
                yr2char = str(yr)[-2:]
                varListAllParcel = ["acres", "acres_AG_" + yr2char,"du_BO","emp_BO",
                                    "netACdu_" + yr2char,"netACemp_" + yr2char]

        #                #make future bikeways layer combining present and proposed bikeways
        #                future_bikeways =  os.path.join(pScratchworkspace, "currAndFutBikewys")     # "in_memory/currAndFutBikewys"
        #                arcpy.Merge_management(["currBikeways_lyr","futBikeways_lyr"],
        #                                        future_bikeways)

                #get unique list of project IDs
                inList = ppautils.getItemList(input_project_lines,projectIdCol,AllItems,specProjList) 
                if(debugLevel & 2)==2: arcpy.AddMessage("{}".format(inList))
                #================ITERATE THROUGH PROJECTS=========================

                if(debugLevel & 2)==2:  
                    sMsg = "nProjectLines = {}, len(inList)={}".format(arcpy.GetCount_management(fl_inprojectslist)[0], len(inList))
                    arcpy.AddMessage(sMsg)
                #arcpy.AddMessage( "-" * 30 + "\nbegin project-level summaries:")
                for i, aProj in enumerate(inList):
                    projID = str(aProj)
                    sMsg = "  Summarizing data for project {} of {} for year {}".format((i + 1), len(inList), yr)
                    arcpy.AddMessage(sMsg)

                    #select project list features with project code of current iteration
                    strSQL = projectIdCol + " = '" + str(aProj) + "'" #note the single quote identifiers


                    arcpy.SelectLayerByAttribute_management(fl_inprojectslist, 
                                                            "NEW_SELECTION", 
                                                            strSQL)

                    pLength = 0
                    pType = []
                    #for project in projCursor:
                    commType = ""
                    with arcpy.da.SearchCursor(fl_inprojectslist, ["Shape@", "type", "ProjectCat"]) as rows:
                        for i,row in enumerate(rows):
                            commType = row[1]
                            if(commType is None):
                                commType = "NA"
                            pShape = row[0]
                            pLength += pShape.length/5280
                            projType = row[2]
                            pType.append(projType) 


                    sType = ""
                    if(len(pType)>0):
                        sType = "".join(str(pType[0])).replace(",",";")

                    stringProjInfo = "{}".format(pLength) #"{},{},{},{}".format(projID,sType, commType, pLength) 


                    #================================ALL-PARCEL CALCULATIONS====================
                    #(inPointlLayer,varList,bufDist = 50, inProjLayer=fl_inprojectslist, 
                            #trim_location = True, bufType="WITHIN_A_DISTANCE"):

                    stringAllParcel = ppautils.pointSum(fl_inAllParcel,varListAllParcel,
                                                    buffer_dist_2,fl_inprojectslist)

                    #================================ILUT CALCULATIONS====================
                    stringILUT = ppautils.pointSum(fl_inilut,varListILUT,buffer_dist_2,
                                                fl_inprojectslist)


                    #=============================INTERSECTION CALCULATIONS=================

                    # get the number of intersections within a buffer
                    arcpy.SelectLayerByLocation_management(fl_inInterSect, 
                                                            "WITHIN_A_DISTANCE",
                                                            fl_inprojectslist, 
                                                            buffer_dist_1)
                    cnt_type1=0
                    cnt_type3=0
                    cnt_type4=0
                    with arcpy.da.SearchCursor(fl_inInterSect, "LINKS")	as aCursor:
                        for aCur in aCursor:
                            linkType = aCur[0]
                            if linkType == 1:
                                cnt_type1 += 1
                            elif linkType == 3:
                                cnt_type3 += 1
                            elif linkType == 4:
                                cnt_type4 += 1
                    #print str(aProj)+"type:"+str(cnt_type3)+","+str(cnt_type4)	
                    stringIntersecn = str(cnt_type1) + "," + str(cnt_type3) + "," + str(cnt_type4)

                    #=====================JOB CENTER CALCULATIONS=================		
                    # Get the number of job centers within 4 mile

                    arcpy.SelectLayerByLocation_management(fl_jobCenter, 
                                                            "WITHIN_A_DISTANCE",
                                                            fl_inprojectslist, 
                                                            buffer_dist_jc)
                    result = arcpy.GetCount_management(fl_jobCenter)
                    stringJobCtr = str(result.getOutput(0))

                    #=================TRANSIT STATS============================
                    stringGTFS = ppautils.transitEvents(fl_inprojectslist,buffer_dist_1,"COUNT_OBJECTID")


                    #===================COLLISION STATS============================
                    #set 50ft buffer so only collisions on/very close to segment are captured

                    stringTIMS = ppautils.collisionSummary(fl_inprojectslist,50,
                                                        varListCollisionSums)

                    #select TAZs that intersect with project centerline segments
                    arcpy.SelectLayerByLocation_management(fl_taz, "INTERSECT",fl_inprojectslist)

                    tazDict = {}
                    
                    #=====================TAZ-LEVEL ACCESSBILITY STATS=========
                    if(debugLevel & 2)==2:  arcpy.AddMessage("access_tbl = {}".format(accessibility_tbl))    
                    taz_acc_fields = ['JOBS30D','JOBS45T','POP','WORKERS30D',
                                      'WORKERS45T','JOBS']
                    with arcpy.da.SearchCursor(accessibility_tbl,"*") as accCur:
                        #create dict of all TAZs along with their accessibility metrics
                        for row in accCur:
                            try:
                                dictvals = []
                                for field in taz_acc_fields:
                                    val = row[accCur.fields.index(field)]
                                    dictvals.append(val)
                                tazDict[row[accCur.fields.index('TAZ')]] = dictvals
                            except:
                                pass   #or give a default value to tazDict[int(i['TAZ'])], however, the problem is that the int(i['TAZ']) could fail too.
                                
                    sumJobsD = 0 #simple sum of jobs within driving n mins driving
                    sumJobsT = 0 ##simple sum of jobs within driving n mins via transit
                    sumWorkersD = 0 #simple sum of workers within driving dist
                    sumWorkersT = 0 #simple sum of workers within transit dist
                    sumProductJobsD = 0 #for each taz, multiply jobacc*pop, then sum those products
                    sumProductJobsT = 0 #same as sumProductJobsD, but for driving
                    sumProductWkrsD = 0 #for each taz, multiple workers accessed * jobs
                    sumProductWkrsT = 0
                    sumTAZjobs = 0
                    sumTAZPops = 0 #sum of pop in all TAZs intersecting project
                    TAZcnt = 0 #count of TAZs the project intersects

                    #create iterator for TAZs that intersect the project segments
                    with arcpy.da.SearchCursor(fl_taz,"TAZ07") as accCursor:
                        for row in accCursor: #get total pop of all tazs intersecting project, plus 
                        #sumproducts for weighted avg.
                            taz = row[0]
                            sumJobsD += tazDict[taz][0]
                            sumJobsT += tazDict[taz][1]
                            sumWorkersD += tazDict[taz][3]
                            sumWorkersT += tazDict[taz][4]
                            sumProductJobsD += tazDict[taz][0]*tazDict[taz][2]
                            sumProductJobsT += tazDict[taz][1]*tazDict[taz][2]
                            sumProductWkrsD += tazDict[taz][3]*tazDict[taz][5]   #driving workers weightd x jobs
                            sumProductWkrsT += tazDict[taz][4]*tazDict[taz][5]   #transit wkrs weighted x jobs
                            sumTAZjobs += tazDict[taz][5]
                            sumTAZPops += tazDict[taz][2]
                            TAZcnt += 1

                    #jobs weighted by resident (POP). If no resident, simple average (total jobs/count of TAZs)
                    if sumTAZPops != 0: 
                        avgPersnJobAccD = str(sumProductJobsD/sumTAZPops)
                        avgPersnJobAccT = str(sumProductJobsT/sumTAZPops)
                    elif sumTAZPops == 0 and TAZcnt > 0:
                        avgPersnJobAccD = str(sumJobsD/TAZcnt)
                        avgPersnJobAccT = str(sumJobsT/TAZcnt)
                    #if the project is not in a TAZ, it is outside of SACOG jurisdiction.
                    #for now am giving job access of zero, but need to know best way to handle these
                    else:
                        avgPersnJobAccD = str(0)
                        avgPersnJobAccT = str(0)

                    #accessibility to workers, weighted by jobs. if no jobs, simple average (tot wkrs/TAZ count)
                    if sumTAZjobs != 0: 
                        avgJobWkrAccD = str(sumProductWkrsD/sumTAZjobs) 
                        avgJobWkrAccT = str(sumProductWkrsT/sumTAZjobs)
                    elif sumTAZPops == 0 and TAZcnt > 0:
                        avgJobWkrAccD = str(sumWorkersD/TAZcnt)
                        avgJobWkrAccT = str(sumWorkersT/TAZcnt)
                    else:
                        avgJobWkrAccD = str(0)
                        avgJobWkrAccT = str(0)

                    stringAccMeasures = avgPersnJobAccD + ',' + avgPersnJobAccT + ',' \
                                        + avgJobWkrAccD + ',' + avgJobWkrAccT

                    #========CREATE TEMPORARY BUFFER POLYGON (PUT AFTER 'SELECT BY LOCATION' SECTIONS)=========

                    #make a 0.25-mile buffer around the selected centerline features (on-project)
                    stringBuffArea1 = ppautils.bufferArea(fl_inprojectslist, "tempbuff", buffer_dist_1)
                    stringBuffArea2 = ppautils.bufferArea(fl_inprojectslist, "tempbuff", buffer_dist_2)

                    stringBuffArea = stringBuffArea1 + ',' + stringBuffArea2

                    #=========================TOTAL NON-FREEWAY CENTERLINE MILES========================

                    #create temporary buffer
                    tmpbuff = os.path.join("in_memory","tmpbuffproj")
                    arcpy.Buffer_analysis(fl_inprojectslist, 
                                                tmpbuff, 
                                                buffer_dist_1, 
                                                "FULL", 
                                                "ROUND", 
                                                "ALL")

                    #get mileage for centerline segments whose centroid is in buffer
                    arcpy.SelectLayerByLocation_management(fl_centerline, 
                                                            "HAVE_THEIR_CENTER_IN",tmpbuff)
                                            
                    totalLength = 0

                    SQL = "CLASS NOT IN ( 'H','RAMP')"
                    with arcpy.da.SearchCursor(fl_centerline,"SHAPE@LENGTH",SQL) as aCursor:
                        for cur in aCursor:
                            shapeLength = cur[0]
                            totalLength += shapeLength

                    stringCLineLength = str(totalLength/5280) #convert feet to miles then to string

                    #====================BIKEWAY CENTERLINE MILES================================
                    #get mileage for bikeway segments whose centroid is in buffer
                    arcpy.SelectLayerByLocation_management(fl_bikeways, "HAVE_THEIR_CENTER_IN",tmpbuff)    
                    class1 = 0
                    class2 = 0
                    class3 = 0
                    class4 = 0

                    with arcpy.da.SearchCursor(fl_bikeways,["SHAPE@LENGTH","BIKE_CLASS"]) as bikeCursor:
                        for aCur in bikeCursor:
                            shapeLength = aCur[0]
                            if aCur[1] == 1:
                                class1 += shapeLength
                            if aCur[1] == 2:
                                class2 += shapeLength
                            if aCur[1] == 3:
                                class3 += shapeLength
                            if aCur[1] == 4:
                                class4 += shapeLength

                    stringBikClsLen = ''
                    lClasses = [class1,class2,class3,class4]
                    for i, dist in enumerate(lClasses):
                        if (i==0):
                            stringBikClsLen = "{}".format(dist/5280) #convert feet to miles
                        else:
                            stringBikClsLen = "{},{}".format(stringBikClsLen, dist/5280)

                    #cnt = 1
                    #for i in [class1,class2,class3,class4]:
                    #    if cnt == 1:
                    #        stringBikClsLen = str(i/5280) #convert feet to miles
                    #    else:
                    #        stringBikClsLen = stringBikClsLen + ',' + str(i/5280)
                    #    cnt += 1

                    #===================MODEL NETWORK CALCULATIONS===============================
                    #capclasses: 1=freeway; 6,26,36=ramp; 56 = aux lane;  8=HOV; 9=HOV connector; 7 = bike path
                    #62/63 = centroid connectors; 99=placeholder/proposed
                    #to get surface street stats, select "not in" the above freeway/hov/ramp values

                    arcpy.SelectLayerByLocation_management(fl_inModelNetwork, "HAVE_THEIR_CENTER_IN", tmpbuff)
                    network_metrx = ["DAYVMT","DAYCVMT","LANEMI","DISTANCE"]

                    SQLFwy = 'CAPCLASS IN (1,6,26,36,56,8,9)'
                    SQLNotFwy = 'CAPCLASS NOT IN (1,6,26,36,56,8,9,7,62,63,99)'

                    stringFwyModBuff = ppautils.modelBuffCalc(fl_inModelNetwork, network_metrx, SQLFwy)
                    stringStreetModBuff = ppautils.modelBuffCalc(fl_inModelNetwork, network_metrx, SQLNotFwy)

                    stringCombModBuff = stringFwyModBuff + ',' + stringStreetModBuff
                    #8/1/2017 - Route miles currently just sum of DISTANCE column. Final calc TBD
                    #===================CONCATENATE OUTPUTS========================
                    sMsg = str(yr) + ',' + propertyValues + ',' +  stringProjInfo + ',' + stringBuffArea + ',' + stringAllParcel + ',' + stringILUT + ',' + stringIntersecn + ',' + stringJobCtr + ',' + stringCLineLength + ','  + stringBikClsLen + ',' + stringGTFS + ',' + stringTIMS + ',' + stringCombModBuff + ','  + stringAccMeasures + '\n'
                    if(debugLevel & 4)==4: arcpy.AddMessage(sMsg) 
                    myfile.write(sMsg)
                    i += 1

        #..myfile.close()
        if(debugLevel & 2) == 2:
            sMsg = "os.path.exists({})={}".format(outputFile, os.path.exists(outputFile))
            arcpy.AddMessage(sMsg)
             
        if(debugLevel & 4)==4:
            with open(outputFile, 'r') as inFile:
                for line in inFile:
                    arcpy.AddMessage(line) 

        #end_time = time.time()
        #elapsed = str(round((end_time - start_time)/60,1))
        #arcpy.AddMessage("Time elapsed: " + elapsed + " mins")
        arcpy.AddMessage("Finished. dt={:0.3f}s".format(time.clock()-dt))

        sOK = C_OK
        pTable = os.path.join("in_memory", "projecteval")
        tableView = "P{}".format(int(time.clock()))
        if(arcpy.Exists(tableView)): arcpy.Delete_management(tableView) 
        arcpy.CopyRows_management(outputFile, pTable)
        arcpy.MakeTableView_management(pTable, tableView) 
        #..Updating the xlsx file by writing the csv contents to it.  make sure the xlsx template (Digest_Summary.xlsx) contains
        #  1. worksheet named '1-import'
        #  2. start row = 5, start column = 1
        #  3. csv contains a line of header (line 1, skipped by the reader), and data (to xlsx) starts from line 2.
        outputFolder = arcpy.env.scratchFolder 
        arcpy.AddMessage("outputFolder={}".format(outputFolder))
        #..Make a copy of the XLSDigest to output folder.
        xlsxName = "XLSDigest_{}_{}.xlsx".format(dProperties.get("Project_ID"),int(time.clock()))
        xlsxFile = os.path.join(outputFolder, xlsxName) 
        #..Get the xmlx template file located at the parent folder of workspace GDB
        xlsxTemplate = os.path.join(os.path.dirname(workspace), C_XLSX_Template) 
        #shutil.copy2(xlsxTemplate, xlsxFile)
        arcpy.AddMessage("template xlsxfile={}\noutput xlsxFile={}".format(xlsxTemplate, xlsxFile))
        arcpy.AddMessage("Does Output Folder Exist? {}".format(os.path.exists(outputFolder)))
        writecsv2xlsx(outputFile, xlsxTemplate, xlsxFile, start_row=4, start_col=0)  
        arcpy.AddMessage("Does Output Excel Sheet Exist? {}".format(os.path.exists(xlsxFile)))		
        tReturn = (sOK, tableView, xlsxFile, outputFile)	
    except:
        sMsg = "{}, {}".format(arcpy.GetMessages(2), trace())
        arcpy.AddWarning(sMsg) 
        sOK = sMsg 
        tReturn = (sOK,)
    finally:
        arcpy.ResetEnvironments()

    return tReturn  

def showMessage(sMsg):
    ssMsg = "you entered: {} datetime.datetime.now()={} time.clock()={}".format(sMsg, datetime.datetime.now(), time.clock())
    #print(ssMsg) 
    arcpy.AddMessage(ssMsg) 
#   D:\10Data\SACOG_PPA\SACOGPPA\PPAData.gdb\Project1Line   1320.0 2640.0 21120 D:\10Data\SACOG_PPA\SACOGPPA\PPAData.gdb
#  D:\10Data\SACOG_PPA\PPA_layers_4ESRI.gdb\Project17Lines 1320.0 2640.0 21120 D:\10Data\SACOG_PPA\PPA_layers_4ESRI.gdb
#  D:\10Data\SACOG_PPA\PPA_layers_4ESRI.gdb\Project3Lines 1320.0 2640.0 21120 D:\10Data\SACOG_PPA\PPA_layers_4ESRI.gdb
#  D:\10Data\SACOG_PPA\PPA_layers_4ESRI.gdb\full_project_list_wCType11162017 1320.0 2640.0 21120 D:\10Data\SACOG_PPA\PPA_layers_4ESRI.gdb
# "D:\10Data\SACOG_PPA\PPA Files4ESRI_01312018\PPA_layers_4ESRI.gdb\full_project_list_wCType11162017" 1320.0 2640.0 21120 "D:\10Data\SACOG_PPA\PPA Files4ESRI_01312018\PPA_layers_4ESRI.gdb"

#  D:\10Data\SACOG_PPA\SACOGPPA\PPAData.gdb\Project1Line 1320 2640 21120 D:\10Data\SACOG_PPA\SACOGPPA\PPAData.gdb 2036 P123 A JurisdictionA 90 30 220
#..Entry points of the module - when the module is used as the starting module.
if(__name__=='__main__'):
    #project file inputs  #...User Input....
    debugLevel = 0
    pScratchGDB = arcpy.env.scratchGDB     #  "in_memory"
    if(debugLevel > 0):  arcpy.AddMessage("PY: {}".format(GetThisFileName()))

    input_project_lines = "full_project_list_wCType11162017" # SHP/GDB file of all projects  UserInput:  LineLayer
    buffer_dist_1 = 1320 #distance in feet for transpo infrastructure buffers
    buffer_dist_2 = 2640 #distance in feet for land-use buffers
    buffer_dist_jc = 21120 # job center buffer distance in feet; 4 mile = 21120ft
    workspace = r"D:\10Data\SACOG_PPA\PPA Files4ESRI_01312018\PPA_layers_4ESRI.gdb"

    input_project_lines = arcpy.GetParameterAsText(0) # SHP/GDB file of all projects  UserInput:  LineLayer
    buffer_dist_1 = arcpy.GetParameterAsText(1)  #1320 #distance in feet for transpo infrastructure buffers
    buffer_dist_2 = arcpy.GetParameterAsText(2)  #2640 #distance in feet for land-use buffers
    buffer_dist_jc = arcpy.GetParameterAsText(3)  #21120 # job center buffer distance in feet; 4 mile = 21120ft
    workspace = arcpy.GetParameterAsText(4) 
#    \\data-svr\Modeling\ProjectLevelPerformanceAssessment\DataLayers_Proof_of_Concept\Online Tool\SACOGPPA_data\SACOGPPA\PPAData.sde'
    year = arcpy.GetParameterAsText(5) 
    project_id = arcpy.GetParameterAsText(6) 
    project_type = arcpy.GetParameterAsText(7) 
    jurisdiction = arcpy.GetParameterAsText(8) 
    pavement_condition_score = arcpy.GetParameterAsText(9) 
    vehicle_speed = arcpy.GetParameterAsText(10) 
    average_daily_traffic_volume = arcpy.GetParameterAsText(11) 
     
    if (not project_type) or (project_type=='#'):  project_type = ""
    
    try:
        year = int(year) 
    except:
        year = -2036 

    try:
        buffer_dist_1 = float(buffer_dist_1)
    except:
        buffer_dist_1 = 1320.0
    try:
        buffer_dist_2 = float(buffer_dist_2)
    except:
        bufferDist2 = 2640.0
    try:
        buffer_dist_jc = float(buffer_dist_jc)
    except:
        buffer_dist_jc = 21120.0
    
    dProperties = dict()
    dProperties.setdefault("Project_ID", project_id) 
    dProperties.setdefault("Project_Type", project_type) 
    dProperties.setdefault("Jurisdiction", jurisdiction)
    try:
        pavement_condition_score = int(pavement_condition_score) 
    except:
        pavement_condition_score = C_MissingValue

    dProperties.setdefault("Pavement_Condition_Score", pavement_condition_score)
    try:
        vehicle_speed = int(vehicle_speed) 
        if(vehicle_speed<0): vehicle_speed=0
    except:
        vehical_speed = 10 
    dProperties.setdefault("Vehicle_speed", vehicle_speed) 

    try:
        average_daily_traffic_volume = int(average_daily_traffic_volume) 
        if(average_daily_traffic_volume<0): average_daily_traffic_volume=0
    except:
        average_daily_traffic_volume = 10 

    dProperties.setdefault("Avg_Daily_Traffic_Vol", average_daily_traffic_volume)
    if(debugLevel & 2)==2: arcpy.AddMessage("dProperties={}".format(dProperties))

    #...Construct the year related datasets, which can later be converted into a years loop, like this:
    dYearDatasets = dict() 
    #for year in [2012,2036]:
    #    parcel = "parcel_{}".format(year) 
    #    intersect = "intersections_{}".format(year)
    #    modelnet = "SM15_modelnet_{}".format(year) 
    #    dYearDatasets.setdefault(year, [parcel, intersect, modelnet]) 

    ILUT2036 =  "parcel_2036"               #  "parcel_{}".format(Year)
    interSect2036 = "intersections_2036"    #  "intersections_{}".format(Year)
    modelNet2036  = "SM15_modelnet_2036"    #  "SM15_modelnet_{}".format(Year)
    access2036 = "access_2036_csv"          #table made from accessbility CSV
    bikewaysFuture = "AllBikeways2036"
    
    lDatasets = list() 
    lDatasets.append(ILUT2036) #0
    lDatasets.append(interSect2036) #1
    lDatasets.append(modelNet2036) #2
    lDatasets.append(access2036) #3
    lDatasets.append(bikewaysFuture) #4
    dYearDatasets.setdefault(2036, lDatasets) #make lDatasets the default key for the 2036 value

    ILUT2012 =  "parcel_2012"               #  "parcel_{}".format(Year)
    interSect2012 = "intersections_2012"    #  "intersections_{}".format(Year)
    modelNet2012  = "SM15_modelnet_2012"     #  "SM15_modelnet_{}".format(Year)
    access2012 = "access_2012_csv" #table made from accessbility CSV
    bikewaysCurrent = "AllBikeways2018"

    lDatasets = list() 
    lDatasets.append(ILUT2012)
    lDatasets.append(interSect2012) 
    lDatasets.append(modelNet2012) 
    lDatasets.append(access2012)
    lDatasets.append(bikewaysCurrent)
    dYearDatasets.setdefault(2012, lDatasets) 
    tReturn = dowork(input_project_lines, buffer_dist_1, buffer_dist_2, buffer_dist_jc, workspace, project_type, dYearDatasets, year, dProperties, pScratchGDB,  debugLevel)
    if(tReturn[0] == C_OK):
        #tableView = tReturn[1]
        xlsxFile = tReturn[2]
        #outputFile = tReturn[3] #returns CSV file, not needed
        arcpy.AddMessage("DEBUG - tReturn: {}".format(tReturn))
        arcpy.AddMessage("DEBUG - xlsxFile: {}".format(xlsxFile))
        #arcpy.AddMessage("DEBUG - outputFile: {}".format(outputFile))
        #arcpy.SetParameterAsText(12, tableView)
        arcpy.SetParameterAsText(12, xlsxFile)
        #arcpy.SetParameterAsText(14, outputFile)



